from flask import Flask, render_template, request, redirect, session
import sqlite3
import os
from werkzeug.security import generate_password_hash, check_password_hash

from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side



BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATABASE = os.path.join(BASE_DIR, "database", "tasks.db")


app = Flask(__name__)
app.secret_key = "supersecretkey"

def init_db():
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL    
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                completed INTEGER DEFAULT 0,
                user_id INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        """)





def get_tasks(user_id):
    with sqlite3.connect(DATABASE) as conn:
        tasks = conn.execute("SELECT * FROM tasks WHERE user_id = ?", (user_id,)).fetchall()
    return tasks

def add_task(title, user_id):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("INSERT INTO tasks (title, user_id) VALUES (?, ?)", (title, user_id))


def delete_task(task_id):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("DELETE FROM tasks WHERE id = ?", (task_id,))

def toggle_task(task_id):
    with sqlite3.connect(DATABASE) as conn:
        conn.execute("""
            UPDATE tasks
            SET completed = NOT completed
            WHERE id = ?
        """, (task_id,))


@app.route("/")
def landing():
    if "user_id" in session:
        return redirect("/home")

    return render_template("landing.html")

@app.route("/home")
def home():
    if "user_id" not in session:
        return redirect("/")

    tasks = get_tasks(session["user_id"])
    username = session["username"]
    return render_template("home.html", tasks=tasks, username=username)


@app.route("/register", methods=["GET", "POST"])
def register():
    if "user_id" in session:
        return redirect("/home")

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        hashed_password = generate_password_hash(password)

        with sqlite3.connect(DATABASE) as conn:
            conn.execute(
                "INSERT INTO users (username, password) VALUES (?, ?)",
                (username, hashed_password)
            )
        return redirect("/login")

    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect("/home")
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        with sqlite3.connect(DATABASE) as conn:
            user = conn.execute(
                "SELECT * FROM users WHERE username = ?",
                (username,)
            ).fetchone()

        if user and check_password_hash(user[2], password):
            session["user_id"] = user[0]
            session["username"] = user[1]
            return redirect("/home")

    return render_template("login.html")

@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect("/login")

@app.route("/add", methods=["POST"])
def add():
    if "user_id" not in session:
        return redirect("/login")

    title = request.form.get("title")

    if title:
        add_task(title, session["user_id"])

    return redirect("/")

@app.route("/delete/<int:task_id>", methods=["POST"])
def delete(task_id):
    delete_task(task_id)
    return redirect("/")

@app.route("/update/<int:task_id>", methods=["POST"])
def update(task_id):
    toggle_task(task_id)
    return redirect("/")

@app.route("/export/pdf")
def export_pdf():
    if "user_id" not in session:
        return redirect("/")
    tasks = get_tasks(session["user_id"])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("My Tasks", styles['Title']))

    for task in tasks:
        status = "✓" if task[2] else "☐"
        text = f"{status} {task[1]}"
        content.append(Paragraph(text, styles['Normal']))

    doc.build(content)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="tasks.pdf",
        mimetype="application/pdf"
    )

@app.route("/export/excell")
def export_excell():
    if "user_id" not in session:
        return redirect("/")
    tasks = get_tasks(session["user_id"])

    # creates an excell file
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.append(["ID", "task", "completed"])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_index, task in enumerate(tasks, start=2):
        status = "Yes" if task[2] else "No"
        ws.append([task[0], task[1], status])

        for col in range(1, 4):
            cell = ws.cell(row=row_index, column=col)
            cell.border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="tasks.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    init_db()
    app.run(debug=True)
